"""WAR-format export: round-trips through the importer, RBAC, CSV."""
import import_war_data
from conftest import auth, make_session


async def _rga(db):
    return await make_session(db, role="level_4", agent_id="RGA_1", email="rga1@test.dev")


async def _add(db, agent_id, office, sales_day, **metrics):
    doc = {
        "entry_id": f"pe_{agent_id}_{sales_day}", "agent_id": agent_id, "office": office,
        "sales_day": sales_day, "sets": 0, "sits": 0, "sales": 0, "ots_sits": 0, "ots_sales": 0,
        "n1": 0, "refs_obtained": 0, "ref_sits": 0, "ref_sales": 0, "pos_sits": 0, "pos_sales": 0,
        "vet_sits": 0, "vet_sales": 0, "gross_alp": 0.0, "net_alp": 0.0,
    }
    doc.update(metrics)
    await db.production_entries.insert_one(doc)


async def test_export_round_trips_through_importer(client, seeded_db):
    await _add(seeded_db, "AG_1", "MCM", "2026-07-01", sales=2, sits=3, n1=1, gross_alp=1000.0)
    await _add(seeded_db, "AG_1", "MCM", "2026-07-02", sales=1, sits=2, gross_alp=500.0)
    await _add(seeded_db, "AG_2", "AMP", "2026-07-01", sales=1, sits=1, gross_alp=300.0)
    token = await _rga(seeded_db)

    r = await client.get("/api/vault/export?start=2026-07-01&end=2026-07-02", headers=auth(token))
    assert r.status_code == 200, r.text
    data = r.json()

    # Feed every performance row back through the importer's make_entry and
    # assert the reconstructed per-agent-per-day metrics match the originals.
    rebuilt = {}
    for tab in data["weekly_tabs"]:
        for perf in tab["performance"]:
            e = import_war_data.make_entry("x", "MCM", tab["date"], perf)
            rebuilt[(tab["date"], perf["agent"])] = e

    a1_d1 = rebuilt[("2026-07-01", "Agent One")]
    assert a1_d1["gross_alp"] == 1000.0 and a1_d1["sales"] == 2 and a1_d1["sits"] == 3 and a1_d1["n1"] == 1
    assert rebuilt[("2026-07-02", "Agent One")]["gross_alp"] == 500.0
    assert rebuilt[("2026-07-01", "Agent Two")]["gross_alp"] == 300.0
    assert data["report_metadata"]["weekly_total_alp"] == 1800.0


async def test_export_defaults_to_current_week(client, seeded_db):
    token = await _rga(seeded_db)
    r = await client.get("/api/vault/export", headers=auth(token))
    assert r.status_code == 200, r.text
    assert "weekly_tabs" in r.json()


async def test_export_requires_level_4(client, seeded_db):
    token = await make_session(seeded_db, role="level_3", agent_id="MGA_1", email="mga1@test.dev")
    r = await client.get("/api/vault/export?start=2026-07-01&end=2026-07-02", headers=auth(token))
    assert r.status_code == 403


async def test_export_csv_has_rows(client, seeded_db):
    await _add(seeded_db, "AG_1", "MCM", "2026-07-01", sales=2, gross_alp=1000.0)
    token = await _rga_admin(seeded_db)   # the spreadsheet is admin-only
    r = await client.get("/api/vault/export?start=2026-07-01&end=2026-07-01&format=csv", headers=auth(token))
    assert r.status_code == 200
    assert "text/csv" in r.headers["content-type"]
    body = r.text
    assert "date,agent,office" in body
    assert "Agent One" in body and "1000" in body


async def test_export_bad_dates_rejected(client, seeded_db):
    token = await _rga(seeded_db)
    assert (await client.get("/api/vault/export?start=07-01-2026&end=2026-07-02", headers=auth(token))).status_code == 400
    assert (await client.get("/api/vault/export?start=2026-07-01", headers=auth(token))).status_code == 400


# ---------------- the agent-by-day spreadsheet is admin-only ----------------

BOOTSTRAP_ADMIN = "linnzi@aoluxor.com"


async def _rga_admin(db):
    """Level 4 AND on EXPORT_EMAILS — the reconciliation exports need both."""
    return await make_session(db, role="level_4", agent_id="RGA_1", email=BOOTSTRAP_ADMIN)


async def test_csv_export_is_refused_to_an_rga_not_on_the_export_list(client, seeded_db):
    """The spreadsheet names every agent and their daily numbers in one file,
    a wider view than the JSON summary an RGA already gets."""
    await _add(seeded_db, "AG_1", "MCM", "2026-07-01", sales=2, gross_alp=1000.0)
    token = await _rga(seeded_db)          # level_4, not on EXPORT_EMAILS
    r = await client.get("/api/vault/export?start=2026-07-01&end=2026-07-01&format=csv",
                         headers=auth(token))
    assert r.status_code == 403


async def test_export_list_is_narrower_than_admin(client, seeded_db):
    """Being an admin must not by itself unlock the reconciliation exports —
    ADMIN_EMAILS carries MJ, EXPORT_EMAILS is meant to carry one person."""
    import server
    assert server.EXPORT_EMAILS < server.ADMIN_EMAILS or \
        server.EXPORT_EMAILS != server.ADMIN_EMAILS
    token = await make_session(seeded_db, role="level_4", agent_id="RGA_1",
                               email="mj@aopremier.com")   # admin, not on EXPORT_EMAILS
    r = await client.get("/api/vault/export?week_start=2026-07-01&format=xlsx",
                         headers=auth(token))
    assert r.status_code == 403


async def test_json_export_still_works_for_a_non_admin_rga(client, seeded_db):
    """Restricting the spreadsheet must not take away the backup export."""
    await _add(seeded_db, "AG_1", "MCM", "2026-07-01", sales=2, gross_alp=1000.0)
    token = await _rga(seeded_db)
    r = await client.get("/api/vault/export?start=2026-07-01&end=2026-07-01",
                         headers=auth(token))
    assert r.status_code == 200


async def test_admin_gets_one_csv_row_per_agent_per_day(client, seeded_db):
    await _add(seeded_db, "AG_1", "MCM", "2026-07-01", sales=2, sits=3, n1=1, gross_alp=1000.0)
    await _add(seeded_db, "AG_1", "MCM", "2026-07-02", sales=1, sits=2, gross_alp=500.0)
    await _add(seeded_db, "AG_2", "AMP", "2026-07-01", sales=1, sits=1, gross_alp=300.0)
    token = await _rga_admin(seeded_db)

    r = await client.get("/api/vault/export?start=2026-07-01&end=2026-07-02&format=csv",
                         headers=auth(token))
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/csv")

    lines = [ln for ln in r.text.strip().splitlines() if ln]
    header = lines[0].split(",")
    assert header[:3] == ["date", "agent", "office"]
    assert "gross_alp" in header and "n1" in header
    assert len(lines) == 4, "header + one row per agent per day"
    # Sorted by date, then descending ALP within the day.
    assert lines[1].startswith("2026-07-01,Agent One,MCM")
    assert lines[2].startswith("2026-07-01,Agent Two,AMP")
    assert lines[3].startswith("2026-07-02,Agent One,MCM")


# ---------------- WAR workbook (.xlsx) ----------------

async def test_xlsx_rebuilds_a_war_workbook_that_reimports(client, seeded_db):
    """The point of the format: a generated file must be readable by the same
    parser that reads the office's real reports, or it is not the same report."""
    import io
    from datetime import date
    import war_import

    await _add(seeded_db, "AG_1", "MCM", "2026-07-01", sets=5, sits=4, sales=3,
               n1=1, gross_alp=802.0)
    token = await _rga_admin(seeded_db)

    r = await client.get("/api/vault/export?week_start=2026-07-01&format=xlsx",
                         headers=auth(token))
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "2026-07-01" in r.headers["content-disposition"]

    parsed = war_import.parse_workbook(io.BytesIO(r.content), date(2026, 7, 1))
    # tabs_found is stripped; the constant keys keep the real reports' trailing
    # spaces ("Mon ", "Tues "), which the generated workbook reproduces.
    assert parsed["tabs_found"] == [t.strip() for t in war_import.TAB_DAY_OFFSET]
    row = parsed["days"]["2026-07-01"][0]
    assert row["name"] == "Agent One"
    assert (row["sets"], row["sits"], row["sales"], row["n1"]) == (5, 4, 3, 1)
    assert row["alp"] == 802


async def test_xlsx_lists_the_whole_roster_even_on_a_quiet_day(client, seeded_db):
    """Real reports carry every agent on every tab. A name vanishing on a quiet
    day is what makes two workbooks impossible to compare side by side."""
    import io
    import openpyxl

    await _add(seeded_db, "AG_1", "MCM", "2026-07-01", sales=1, gross_alp=100.0)
    token = await _rga_admin(seeded_db)
    r = await client.get("/api/vault/export?week_start=2026-07-01&format=xlsx",
                         headers=auth(token))
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)

    # Friday: nobody produced, but the roster is still listed.
    names = [row[5] for row in wb["Fri"].iter_rows(min_row=5, values_only=True) if row[5]]
    assert "Agent One" in names
    assert all(v is None for v in list(wb["Fri"].iter_rows(min_row=5, values_only=True))[0][6:20])


async def test_xlsx_spans_nine_days_like_a_real_report(client, seeded_db):
    """Wed (2)/Thurs (2) reach into the next week, which is what creates the
    two-day overlap the whole import rule turns on."""
    import io
    from datetime import date
    import war_import

    await _add(seeded_db, "AG_1", "MCM", "2026-07-09", sales=2, gross_alp=900.0)  # Wed (2)
    token = await _rga_admin(seeded_db)
    r = await client.get("/api/vault/export?week_start=2026-07-01&format=xlsx",
                         headers=auth(token))
    parsed = war_import.parse_workbook(io.BytesIO(r.content), date(2026, 7, 1))
    assert "2026-07-09" in parsed["days"], "the 8th day must land on the Wed (2) tab"


async def test_xlsx_requires_a_week_start(client, seeded_db):
    token = await _rga_admin(seeded_db)
    r = await client.get("/api/vault/export?start=2026-07-01&end=2026-07-02&format=xlsx",
                         headers=auth(token))
    assert r.status_code == 400


async def test_unknown_format_is_rejected(client, seeded_db):
    token = await _rga_admin(seeded_db)
    r = await client.get("/api/vault/export?week_start=2026-07-01&format=pdf",
                         headers=auth(token))
    assert r.status_code == 400
