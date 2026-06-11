"""
XLSX WAR Import Script
Reads xlsx WAR spreadsheets and imports daily agent production into MongoDB.

Run locally (not from the cloud container — MongoDB requires local network):
    pip install openpyxl pymongo dnspython
    python import_xlsx_war.py /path/to/file1.xlsx /path/to/file2.xlsx ...

Or drop xlsx files in backend/data/xlsx_war/ and run without args:
    python import_xlsx_war.py
"""
import re
import sys
import os
import uuid
from datetime import datetime, timezone, timedelta, date

import openpyxl
from pymongo import MongoClient

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017/")
DB_NAME = os.environ.get("MONGO_DB", "vantagelife")

DETROIT_OFFSET = timedelta(hours=-4)  # EDT

# Tabs that represent daily production, mapped to day offset from week start (Wednesday)
TAB_DAY_OFFSET = {
    "Wed": 0,
    "Thurs": 1,
    "Fri": 2,
    "Sat": 3,
    "Sun": 4,
    "Mon ": 5,
    "Tues ": 6,
    "Wed (2)": 7,
    "Thurs (2)": 8,
}

# Week start date — override via WEEK_START env var (YYYY-MM-DD) or defaults to 2026-05-06
try:
    WEEK_START = date.fromisoformat(os.environ["WEEK_START"])
except (KeyError, ValueError):
    WEEK_START = date(2026, 5, 6)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def detroit_submit_time(date_str: str) -> datetime:
    d = datetime.strptime(date_str, "%Y-%m-%d")
    local = d.replace(hour=20, minute=30, tzinfo=timezone(DETROIT_OFFSET))
    return local.astimezone(timezone.utc)


def safe_int(v) -> int:
    if v is None:
        return 0
    try:
        f = float(v)
        return 0 if f != f else int(f)  # NaN → 0
    except (ValueError, TypeError):
        return 0


def safe_float(v) -> float:
    if v is None:
        return 0.0
    try:
        f = float(v)
        return 0.0 if f != f else f
    except (ValueError, TypeError):
        return 0.0


def has_activity(row: dict) -> bool:
    """True if the agent recorded at least one non-zero metric."""
    keys = ("sets", "sits", "sales", "ots_sits", "ots_sales", "n1",
            "refs_obtained", "ref_sits", "ref_sales",
            "pos_sits", "pos_sales", "vet_sits", "vet_sales", "alp")
    return any(safe_float(row.get(k)) != 0.0 for k in keys)


def get_or_create_agent(db, name: str, office: str) -> str:
    existing = db.agent_profiles.find_one(
        {"name": {"$regex": f"^{re.escape(name.strip())}$", "$options": "i"}}
    )
    if existing:
        return existing["agent_id"]
    agent_id = f"agent_{uuid.uuid4().hex[:10]}"
    db.agent_profiles.insert_one({
        "agent_id": agent_id,
        "name": name.strip(),
        "office": office,
        "role": "level_1",
        "upline_id": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    print(f"    + Created agent: {name} ({agent_id})")
    return agent_id


def make_entry(agent_id: str, office: str, date_str: str, m: dict) -> dict:
    gross_alp = safe_float(m.get("alp"))
    return {
        "entry_id": f"pe_{uuid.uuid4().hex[:12]}",
        "agent_id": agent_id,
        "office": office,
        "sales_day": date_str,
        "sets": safe_int(m.get("sets")),
        "sits": safe_int(m.get("sits")),
        "sales": safe_int(m.get("sales")),
        "ots_sits": safe_int(m.get("ots_sits")),
        "ots_sales": safe_int(m.get("ots_sales")),
        "n1": safe_int(m.get("n1")),
        "refs_obtained": safe_int(m.get("refs_obtained")),
        "ref_sits": safe_int(m.get("ref_sits")),
        "ref_sales": safe_int(m.get("ref_sales")),
        "pos_sits": safe_int(m.get("pos_sits")),
        "pos_sales": safe_int(m.get("pos_sales")),
        "vet_sits": safe_int(m.get("vet_sits")),
        "vet_sales": safe_int(m.get("vet_sales")),
        "gross_alp": gross_alp,
        "net_alp": gross_alp,
        "submitted_at": detroit_submit_time(date_str),
        "submitted_on_time": True,
        "is_adjustment": False,
        "source": "war_xlsx_import",
    }


# ---------------------------------------------------------------------------
# xlsx parsing
# ---------------------------------------------------------------------------

_KEEP_UPPER = {"RGA", "MGA", "GA"}


def extract_office_name(ws) -> str:
    """Read the RGA office name from the summary header (row index 1, col index 1)."""
    rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    if not rows or len(rows[0]) < 2:
        return "Unknown"
    raw = str(rows[0][1] or "Unknown")
    # Normalise ALL-CAPS strings ("RUST RGA" → "Rust RGA"), preserve mixed-case as-is
    if raw == raw.upper():
        words = [w if w in _KEEP_UPPER else w.capitalize() for w in raw.split()]
        return " ".join(words)
    return raw


def parse_daily_tab(ws) -> list[dict]:
    """
    Locate the data section (starts with MGA/GA/SA/LDR header row) and extract
    agent rows that have at least one non-zero metric.

    Column layout in the data section:
      0: MGA  1: GA  2: SA  3: LDR  4: State  5: Agent
      6: SETS  7: SITS  8: SALES  9: OTS SITS  10: OTS SALES  11: N1
      12: REF OBT  13: REF SITS  14: REF SALES
      15: POS SITS  16: POS SALES  17: VET SITS  18: VET SALES  19: ALP
    """
    results = []
    in_data = False

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        # Detect the data section header row
        if not in_data:
            if len(row) >= 7 and row[0] == "MGA" and row[5] == "Agent" and row[6] == "SETS":
                in_data = True
            continue

        if len(row) < 20:
            continue

        agent_name = row[5]
        if not agent_name or not str(agent_name).strip():
            continue
        agent_name = str(agent_name).strip()

        # Skip formula remnants (shouldn't appear in data section but guard anyway)
        if agent_name.startswith("="):
            continue

        m = {
            "name": agent_name,
            "sets": row[6],
            "sits": row[7],
            "sales": row[8],
            "ots_sits": row[9],
            "ots_sales": row[10],
            "n1": row[11],
            "refs_obtained": row[12],
            "ref_sits": row[13],
            "ref_sales": row[14],
            "pos_sits": row[15],
            "pos_sales": row[16],
            "vet_sits": row[17],
            "vet_sales": row[18],
            "alp": row[19],
        }

        if has_activity(m):
            results.append(m)

    return results


def import_xlsx_file(db, path: str, week_start: date) -> int:
    print(f"\nProcessing {os.path.basename(path)} ...")
    wb = openpyxl.load_workbook(path, data_only=True)

    # Get office name from any daily tab (they all share the same header)
    office = extract_office_name(wb["Wed"])
    print(f"  Office: {office}")

    total = 0
    for tab_name, day_offset in TAB_DAY_OFFSET.items():
        if tab_name not in wb.sheetnames:
            continue

        date_str = (week_start + timedelta(days=day_offset)).strftime("%Y-%m-%d")
        ws = wb[tab_name]
        agents = parse_daily_tab(ws)

        if not agents:
            continue

        print(f"  {tab_name} ({date_str}): {len(agents)} agents with activity")
        for a in agents:
            name = a["name"]
            agent_id = get_or_create_agent(db, name, office)

            if db.production_entries.find_one({"agent_id": agent_id, "sales_day": date_str}):
                print(f"    ~ Skipping duplicate: {name} on {date_str}")
                continue

            entry = make_entry(agent_id, office, date_str, a)
            db.production_entries.insert_one(entry)
            total += 1
            print(f"    [OK] {name} | ALP ${entry['gross_alp']:,.0f}")

    return total


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) > 1:
        files = sys.argv[1:]
    else:
        xlsx_dir = os.path.join(os.path.dirname(__file__), "data", "xlsx_war")
        if os.path.isdir(xlsx_dir):
            files = [
                os.path.join(xlsx_dir, f)
                for f in sorted(os.listdir(xlsx_dir))
                if f.endswith(".xlsx")
            ]
        else:
            print("Usage: python import_xlsx_war.py file1.xlsx file2.xlsx ...")
            print("  Or place files in backend/data/xlsx_war/ and run without args.")
            sys.exit(1)

    if not files:
        print("No xlsx files found.")
        sys.exit(1)

    print("Connecting to MongoDB ...")
    client = MongoClient(MONGO_URL, serverSelectionTimeoutMS=15000)
    client.admin.command("ping")
    db = client[DB_NAME]
    print(f"Connected to '{DB_NAME}'\n")

    grand_total = 0
    for f in files:
        count = import_xlsx_file(db, f, WEEK_START)
        print(f"  → {count} entries inserted")
        grand_total += count

    print(f"\nDone. Total entries inserted: {grand_total}")
    client.close()


if __name__ == "__main__":
    main()
