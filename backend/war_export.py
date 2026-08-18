"""Write a WAR-format .xlsx from the app's own production data.

Once agents enter their numbers nightly in the app, the office still wants the
workbook it has always read. This rebuilds it: same tabs, same column layout,
same header row as the spreadsheets the importer reads — so a generated file can
be opened side by side with an old WAR report, and re-imported unchanged.

Every structural constant comes from war_import. Defining the layout twice is
how a column shift silently corrupts one path but not the other, so this module
owns none of it.
"""
from __future__ import annotations

import io
from datetime import date, timedelta
from typing import Any, Dict, List, Optional

import openpyxl

import metrics
import war_import


def _agent_row(person: Dict[str, Any], m: Optional[Dict[str, Any]]) -> List[Any]:
    """One data row: leadership columns, then the 14 metrics, then Close Rate.

    `m` is None for an agent with nothing that day. Real reports leave those
    cells blank rather than zero, and the importer's activity filter drops
    either, so blank keeps the generated file closest to the original.
    """
    row: List[Any] = [None] * (war_import.SHOW_RATE_COLUMN + 1)
    for key, col in war_import.CONTEXT_COLUMNS.items():
        row[col] = person.get(key)
    row[war_import.AGENT_NAME_COLUMN] = person["name"]
    if m is None:
        return row
    for key, col in war_import.METRIC_COLUMNS.items():
        row[col] = m.get(key, 0)
    row[war_import.CLOSE_RATE_COLUMN] = round(
        metrics.close_rate(int(m.get("sales") or 0), int(m.get("sits") or 0)), 1)
    # Show Rate is left blank on purpose: it is a real WAR column but its
    # formula has never been specified for this app, and inventing one would
    # put a fabricated number in a report someone reconciles against.
    return row


def _write_tab(wb, title: str, office: str, roster: List[Dict[str, Any]],
               rows_by_agent: Dict[str, Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    total_alp = sum(float(m.get("alp") or 0) for m in rows_by_agent.values())
    # Summary block above the data section, matching the real layout: the
    # office name sits at row 2 column 2, which is where extract_office_name
    # reads it back from.
    ws.append([None, "ALP", None])
    ws.append([None, office, round(total_alp, 2)])
    ws.append([None, None, None])
    ws.append(list(war_import.HEADER_ROW))
    for person in roster:
        ws.append(_agent_row(person, rows_by_agent.get(person["name"])))


def build_workbook(office: str, week_start: date, roster: List[Dict[str, Any]],
                   by_day: Dict[str, Dict[str, Dict[str, Any]]]) -> io.BytesIO:
    """Build the workbook.

    roster:  [{"name", "mga", "ga", "sa", "state"}, ...] — every agent in scope,
             listed on every tab whether or not they produced, as the real
             reports do.
    by_day:  {"YYYY-MM-DD": {agent_name: {metric: value, ..., "alp": float}}}

    Tabs follow war_import.TAB_DAY_OFFSET, so the file spans the same nine days
    as a real report and "Wed (2)"/"Thurs (2)" line up with the next week.
    Weekly Totals sums the seven days of this week only.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    totals: Dict[str, Dict[str, Any]] = {}
    for offset in range(7):
        day = (week_start + timedelta(days=offset)).strftime("%Y-%m-%d")
        for name, m in by_day.get(day, {}).items():
            acc = totals.setdefault(name, {k: 0 for k in war_import.METRIC_COLUMNS})
            for key in war_import.METRIC_COLUMNS:
                acc[key] = (acc[key] or 0) + float(m.get(key) or 0)
    _write_tab(wb, war_import.TOTALS_TAB, office, roster, totals)

    for tab, offset in war_import.TAB_DAY_OFFSET.items():
        day = (week_start + timedelta(days=offset)).strftime("%Y-%m-%d")
        _write_tab(wb, tab, office, roster, by_day.get(day, {}))

    # Empty but present, so the generated file has the same tab set as a real
    # report. The app does not track lost business; writing rows here would be
    # inventing them.
    wb.create_sheet(war_import.LOST_BUSINESS_TAB).append(list(war_import.HEADER_ROW))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
