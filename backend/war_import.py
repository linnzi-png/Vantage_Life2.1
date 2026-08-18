"""
Shared WAR spreadsheet parsing.

Pure parsing only — no database access — so the CLI backfill script
(import_xlsx_war.py) and the in-app admin upload endpoint
(POST /api/admin/import-war-report) read the exact same column layout.
Keep every structural constant here; duplicating the layout in two
places is how a column shift silently corrupts one path but not the other.
"""
from datetime import datetime, timezone, date, timedelta
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo
import uuid

import openpyxl

# Tabs that represent daily production, mapped to day offset from week start
# (the week start is the Wednesday the report is named for).
#
# Note the report covers NINE days, not seven: "Wed (2)" and "Thurs (2)" are
# the same calendar days as the NEXT report's "Wed" and "Thurs". Consecutive
# weekly files therefore overlap by two days — see resolve conflict handling
# in the importer callers, which treat the later file as authoritative.
TAB_DAY_OFFSET = {
    "Wed": 0,
    "Thurs": 1,
    "Fri": 2,
    "Sat": 3,
    "Sun": 4,
    "Mon ": 5,
    "Tues ": 6,
    "Wed (2)": 7,
    "Thurs (2)": 8,
}

# The 14 nightly metrics, in the canonical order, mapped to their column index
# in the daily-tab data section. Column layout:
#   0: MGA  1: GA  2: SA  3: LDR  4: State  5: Agent
#   6: SETS  7: SITS  8: SALES  9: OTS SITS  10: OTS SALES  11: N1
#   12: REF OBT  13: REF SITS  14: REF SALES
#   15: POS SITS  16: POS SALES  17: VET SITS  18: VET SALES  19: ALP
METRIC_COLUMNS = {
    "sets": 6,
    "sits": 7,
    "sales": 8,
    "ots_sits": 9,
    "ots_sales": 10,
    "n1": 11,
    "refs_obtained": 12,
    "ref_sits": 13,
    "ref_sales": 14,
    "pos_sits": 15,
    "pos_sales": 16,
    "vet_sits": 17,
    "vet_sales": 18,
    "alp": 19,
}

AGENT_NAME_COLUMN = 5

# Trailing computed columns. Present so a generated workbook has the same shape
# as a real one; the parser ignores them and recomputes from the raw metrics.
CLOSE_RATE_COLUMN = 20
SHOW_RATE_COLUMN = 21

# The data-section header, exactly as the real reports write it. parse_daily_tab
# finds the data section by matching row[0]/row[5]/row[6] against it, so the
# writer and the reader must share this one definition or a generated workbook
# would not re-import.
HEADER_ROW = [
    "MGA", "GA", "SA", "LDR", "RESIDENT \nSTATE", "Agent",
    "SETS", "SITS", "SALES", "OTS SITS", "OTS SALES", "N1",
    "REF OBT", "REF SITS", "REF SALES", "POS SITS", "POS SALES",
    "VETS SITS", "VET SALES", "ALP", "Close Rate", "Show Rate",
]

# Row the office summary block puts the office name on (1-based, column 2),
# and the row the data-section header lands on beneath it.
OFFICE_NAME_ROW = 2
HEADER_ROW_INDEX = 4

TOTALS_TAB = "Weekly Totals"

# Present in every real report. Not a daily tab and never parsed — the app has
# no lost-business tracking, so a generated workbook writes it empty rather than
# omit it (shape matches) or invent rows (it would be fabricated data).
LOST_BUSINESS_TAB = "LOST BUSINESS"

# Leadership columns. Not imported as data — carried alongside unmatched agent
# names so an admin can see who a missing agent reports to and onboard them
# through the existing Add Person flow with the right upline.
CONTEXT_COLUMNS = {"mga": 0, "ga": 1, "sa": 2, "state": 4}

# Marks entries written by a WAR import. The upload endpoint will only
# overwrite entries carrying this source, so an agent's own in-app submission
# is never silently replaced by a spreadsheet backfill.
WAR_IMPORT_SOURCE = "war_xlsx_import"

# Integer metrics are every metric except gross ALP, which is a dollar amount.
_INT_METRICS = [k for k in METRIC_COLUMNS if k != "alp"]

_KEEP_UPPER = {"RGA", "MGA", "GA"}


# ---------------------------------------------------------------------------
# coercion helpers
# ---------------------------------------------------------------------------

def safe_int(v: Any) -> int:
    if v is None:
        return 0
    try:
        f = float(v)
        return 0 if f != f else int(f)  # NaN → 0
    except (ValueError, TypeError):
        return 0


def safe_float(v: Any) -> float:
    if v is None:
        return 0.0
    try:
        f = float(v)
        return 0.0 if f != f else f
    except (ValueError, TypeError):
        return 0.0


def has_activity(row: Dict[str, Any]) -> bool:
    """True if the agent recorded at least one non-zero metric.

    Blank rows are the norm: every roster agent appears on every daily tab
    whether or not they worked, so without this filter a 29-agent office
    would write 29 zero-entries per day.
    """
    return any(safe_float(row.get(k)) != 0.0 for k in METRIC_COLUMNS)


def detroit_submit_time(date_str: str) -> datetime:
    """Backfilled entries are stamped 8:30 PM Detroit — inside the normal
    nightly window, so they never look like Midnight Miracle submissions."""
    d = datetime.strptime(date_str, "%Y-%m-%d")
    local = d.replace(hour=20, minute=30, tzinfo=ZoneInfo("America/Detroit"))
    return local.astimezone(timezone.utc)


# ---------------------------------------------------------------------------
# xlsx parsing
# ---------------------------------------------------------------------------

def extract_office_name(ws) -> str:
    """Read the RGA office name from the summary header (row 2, column 2)."""
    rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    if not rows or len(rows[0]) < 2:
        return "Unknown"
    raw = str(rows[0][1] or "Unknown")
    # Normalise ALL-CAPS strings ("RUST RGA" → "Rust RGA"), preserve mixed-case as-is
    if raw == raw.upper():
        words = [w if w in _KEEP_UPPER else w.capitalize() for w in raw.split()]
        return " ".join(words)
    return raw


def parse_daily_tab(ws) -> List[Dict[str, Any]]:
    """Extract agent rows with at least one non-zero metric from one daily tab."""
    return parse_daily_tab_split(ws)[0]


def parse_daily_tab_split(ws) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Split one daily tab into (rows with activity, names reported as nothing).

    The second list is what the activity filter throws away, and it carries real
    meaning. A row that is present but blank or zeroed is the office stating
    that this agent produced nothing that day. Without those names the importer
    cannot tell "not mentioned in this file" from "reported as nothing", so a
    retraction in a later report can never clear the earlier report's value and
    the superseded number survives forever.

    The data section begins at the header row whose first cell is "MGA" —
    everything above it is the office summary block, which repeats the same
    metric names and would otherwise be misread as agent rows.
    """
    active: List[Dict[str, Any]] = []
    silent: List[str] = []
    in_data = False
    max_col = max(METRIC_COLUMNS.values())

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        # Detect the data section header row
        if not in_data:
            if (len(row) >= 7 and row[0] == "MGA"
                    and row[AGENT_NAME_COLUMN] == "Agent" and row[6] == "SETS"):
                in_data = True
            continue

        if len(row) <= max_col:
            continue

        agent_name = row[AGENT_NAME_COLUMN]
        if not agent_name or not str(agent_name).strip():
            continue
        agent_name = str(agent_name).strip()

        # Skip formula remnants (shouldn't appear in the data section, but guard)
        if agent_name.startswith("="):
            continue

        m: Dict[str, Any] = {"name": agent_name}
        for key, col in METRIC_COLUMNS.items():
            m[key] = row[col]
        for key, col in CONTEXT_COLUMNS.items():
            v = row[col] if col < len(row) else None
            m[key] = str(v).strip() if v else None

        if has_activity(m):
            active.append(m)
        else:
            silent.append(agent_name)

    return active, silent


def parse_workbook(file_obj, week_start: date) -> Dict[str, Any]:
    """Parse a whole WAR workbook into office name + dated agent rows.

    `file_obj` may be a path or any binary file-like object (an uploaded
    file stream). Returns:
        {"office": str,
         "days":   {"YYYY-MM-DD": [row, ...]},   # rows with activity
         "silent": {"YYYY-MM-DD": [name, ...]},  # rows reported as nothing
         "tabs_found": [...]}

    "silent" is populated for every tab present, including tabs with no active
    rows at all. Callers decide which days this file is authoritative for —
    see the supersede guard in the import endpoint.
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    try:
        # All daily tabs share the same summary header; fall back to the first
        # sheet only if the canonical "Wed" tab is missing.
        header_sheet = wb["Wed"] if "Wed" in wb.sheetnames else wb[wb.sheetnames[0]]
        office = extract_office_name(header_sheet)

        # Stripped→actual sheet name map so "Mon" and "Mon " both match.
        sheet_map = {s.strip(): s for s in wb.sheetnames}

        days: Dict[str, List[Dict[str, Any]]] = {}
        silent: Dict[str, List[str]] = {}
        tabs_found: List[str] = []
        for tab_name, day_offset in TAB_DAY_OFFSET.items():
            actual_name = sheet_map.get(tab_name.strip())
            if actual_name is None:
                continue
            tabs_found.append(tab_name.strip())
            rows, quiet = parse_daily_tab_split(wb[actual_name])
            date_str = (week_start + timedelta(days=day_offset)).strftime("%Y-%m-%d")
            if quiet:
                silent[date_str] = quiet
            if not rows:
                continue
            days[date_str] = rows

        return {"office": office, "days": days, "silent": silent,
                "tabs_found": tabs_found}
    finally:
        wb.close()


def week_start_from_filename(filename: str) -> Optional[date]:
    """WAR files are named '<YYYY-MM-DD>_<Office>_War_Report.xlsx', where the
    date is the Wednesday the report opens on. Returns None if unparseable."""
    stem = filename.replace("\\", "/").rsplit("/", 1)[-1]
    try:
        return date.fromisoformat(stem[:10])
    except ValueError:
        return None


def build_entry(agent_id: str, office: str, date_str: str, m: Dict[str, Any]) -> Dict[str, Any]:
    """Build a production_entries document from one parsed agent row."""
    gross_alp = safe_float(m.get("alp"))
    entry = {
        "entry_id": f"pe_{uuid.uuid4().hex[:12]}",
        "agent_id": agent_id,
        "office": office,
        "sales_day": date_str,
        "gross_alp": gross_alp,
        "net_alp": gross_alp,
        "submitted_at": detroit_submit_time(date_str),
        "submitted_on_time": True,
        "is_adjustment": False,
        "source": WAR_IMPORT_SOURCE,
    }
    for key in _INT_METRICS:
        entry[key] = safe_int(m.get(key))
    return entry
